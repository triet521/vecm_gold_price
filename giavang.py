from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# ⚙️ Cấu hình Selenium
service = Service(r"C:\Users\Me\Downloads\Compressed\edgedriver_win64\msedgedriver.exe")
options = Options()
options.add_argument("--headless=new")
driver = webdriver.Edge(service=service, options=options)
driver.get("https://giavang.net/")

# ⏳ Chờ dữ liệu JS load xong
WebDriverWait(driver, 15).until(
    EC.presence_of_element_located((By.CSS_SELECTOR, "tr[data-code] td[data-field=buy] span#text"))
)

# ✅ Lấy dữ liệu
rows = driver.find_elements(
    By.CSS_SELECTOR,
    "table#tbl tr.bg-light-yellow, table#tbl tr.bg-white"
)

scrape_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
data = []

for row in rows:
    try:
        name = row.find_element(By.CSS_SELECTOR, "td:nth-child(1)").text.strip()
        mua = row.find_element(By.CSS_SELECTOR, "td[data-field=buy] span#text").text
        ban = row.find_element(By.CSS_SELECTOR, "td[data-field=sell] span#text").text

        m = int(mua.replace(".", ""))
        b = int(ban.replace(".", ""))
        if m > 1_000_000 and b > 1_000_000:
            data.append([scrape_time, name, m, b])
            print(f"✔️ {scrape_time} – {name}: {m:,} – {b:,}")
    except:
        continue

driver.quit()

# 📁 File Excel lưu lịch sử
file_path = "giavang.xlsx"

# 📤 Mở hoặc tạo mới file Excel
if os.path.exists(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Giá vàng hôm nay"
    # Header
    ws.append(["Thời gian", "Loại vàng", "Mua vào", "Bán ra"])
    for col in range(1, 5):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4F81BD")
        c.alignment = Alignment("center")

# 📝 Ghi thêm dữ liệu mới
for ts, name, m, b in data:
    ws.append([ts, name, m, b])

# 📐 Căn chỉnh cột (chỉ làm 1 lần cho file mới)
if ws.max_row == len(data) + 1:  # tức là vừa tạo file mới
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

wb.save(file_path)
print("✅ Đã ghi dữ liệu lịch sử vào giavang.xlsx.")
